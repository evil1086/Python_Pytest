import openpyxl


class HomePageData:

    test_HomePage_data = [{"firstname":"A","lastname":"B","gender":"Male"}, {"firstname":"A", "lastname":"B", "gender":"Female"}]
    test_login_details = {"username":"standard_user", "password": "secret_sauce", "fname": "vinod", "lname": "Pawar", "Zip": "873564"}

    @staticmethod
    def getTestData(test_case_name):
        Dict = {}
        book = openpyxl.load_workbook("C:\\Users\\Owner\\Documents\\PythonDemo.xlsx")
        sheet = book.active
        for i in range(1, sheet.max_row + 1):  # to get rows
            if sheet.cell(row=i, column=1).value == test_case_name:

                for j in range(2, sheet.max_column + 1):  # to get columns
                    # Dict["lastname"]="shetty
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        return[Dict]
    
    

